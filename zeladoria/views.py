from io import BytesIO

from django.contrib import messages
from django.db.models import Count, Q
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .forms import ChamadoCreateForm, ChamadoFollowUpForm
from .models import ChamadoZeladoria


def index(request):
    total_abertos = ChamadoZeladoria.objects.exclude(
        status__in=[ChamadoZeladoria.Status.CONCLUIDO, ChamadoZeladoria.Status.CANCELADO]
    ).count()
    total_mes = ChamadoZeladoria.objects.filter(
        criado_em__year=timezone.localdate().year,
        criado_em__month=timezone.localdate().month,
    ).count()
    aguardando_ticket = ChamadoZeladoria.objects.filter(ticket_oficial='').exclude(
        status__in=[ChamadoZeladoria.Status.CONCLUIDO, ChamadoZeladoria.Status.CANCELADO]
    ).count()

    return render(
        request,
        'zeladoria/index.html',
        {
            'total_abertos': total_abertos,
            'total_mes': total_mes,
            'aguardando_ticket': aguardando_ticket,
        },
    )


def criar(request):
    if request.method == 'POST':
        form = ChamadoCreateForm(request.POST, request.FILES)
        if form.is_valid():
            chamado = form.save()
            messages.success(request, 'Necessidade cadastrada com sucesso.')
            return redirect('zeladoria:detalhe', pk=chamado.pk)
    else:
        form = ChamadoCreateForm()

    return render(request, 'zeladoria/form.html', {'form': form})


def painel(request):
    status = request.GET.get('status', '')
    busca = request.GET.get('busca', '').strip()
    mes = request.GET.get('mes', '')

    chamados = ChamadoZeladoria.objects.all()

    if status:
        chamados = chamados.filter(status=status)

    if busca:
        chamados = chamados.filter(
            Q(solicitante__icontains=busca)
            | Q(titulo__icontains=busca)
            | Q(local__icontains=busca)
            | Q(descricao__icontains=busca)
            | Q(ticket_oficial__icontains=busca)
        )

    if mes:
        try:
            ano, numero_mes = mes.split('-')
            chamados = chamados.filter(criado_em__year=ano, criado_em__month=numero_mes)
        except ValueError:
            messages.error(request, 'Filtro de mes invalido.')

    resumo_status = ChamadoZeladoria.objects.values('status').annotate(total=Count('id'))

    return render(
        request,
        'zeladoria/painel.html',
        {
            'chamados': chamados,
            'status_choices': ChamadoZeladoria.Status.choices,
            'status_atual': status,
            'busca': busca,
            'mes': mes,
            'resumo_status': resumo_status,
        },
    )


def detalhe(request, pk):
    chamado = get_object_or_404(ChamadoZeladoria, pk=pk)
    return render(request, 'zeladoria/detalhe.html', {'chamado': chamado})


def editar(request, pk):
    chamado = get_object_or_404(ChamadoZeladoria, pk=pk)

    if request.method == 'POST':
        form = ChamadoFollowUpForm(request.POST, instance=chamado)
        if form.is_valid():
            form.save()
            messages.success(request, 'Follow-up atualizado com sucesso.')
            return redirect('zeladoria:painel')
    else:
        form = ChamadoFollowUpForm(instance=chamado)

    return render(request, 'zeladoria/editar.html', {'form': form, 'chamado': chamado})


def exportar_excel(request):
    chamados = ChamadoZeladoria.objects.all().order_by('-criado_em')
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'Chamados'

    title = 'Relatorio de Zeladoria Predial'
    sheet.merge_cells('A1:K1')
    sheet['A1'] = title
    sheet['A1'].font = Font(bold=True, size=16, color='FFFFFF')
    sheet['A1'].fill = PatternFill('solid', fgColor='1967D2')
    sheet['A1'].alignment = Alignment(horizontal='center')

    headers = [
        'ID',
        'Titulo',
        'Mes',
        'Data',
        'Solicitante',
        'Local',
        'Descricao',
        'Status',
        'Ticket oficial',
        'Observacoes',
        'Foto',
    ]
    sheet.append(headers)

    header_fill = PatternFill('solid', fgColor='E8F3EF')
    for cell in sheet[2]:
        cell.font = Font(bold=True, color='17212B')
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    for chamado in chamados:
        sheet.append(
            [
                chamado.id,
                chamado.titulo,
                chamado.mes_referencia,
                timezone.localtime(chamado.criado_em).strftime('%d/%m/%Y %H:%M'),
                chamado.solicitante,
                chamado.local,
                chamado.descricao,
                chamado.get_status_display(),
                chamado.ticket_oficial,
                chamado.observacoes,
                request.build_absolute_uri(chamado.foto.url) if chamado.foto else '',
            ]
        )

    for row in sheet.iter_rows(min_row=3):
        for cell in row:
            cell.alignment = Alignment(vertical='top', wrap_text=True)

    widths = [8, 28, 12, 18, 24, 28, 45, 18, 18, 38, 42]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width

    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    filename = f"relatorio_zeladoria_{timezone.localdate().strftime('%Y_%m_%d')}.xlsx"
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response

# Create your views here.
