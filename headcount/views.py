from datetime import datetime
from io import BytesIO

from django.contrib import messages
from django.db.models import Count
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .forms import HeadcountUploadForm
from .models import BirthdayList, HeadcountImport
from .services import compare_with_previous_month, import_headcount, process_birthday_list


def index(request):
    ultima_importacao = HeadcountImport.objects.first()
    total_atual = ultima_importacao.total_membros if ultima_importacao else 0
    aniversariantes_mes = BirthdayList.objects.first()
    total_aniversariantes = aniversariantes_mes.nomes.count() if aniversariantes_mes else 0

    return render(
        request,
        'headcount/index.html',
        {
            'ultima_importacao': ultima_importacao,
            'total_atual': total_atual,
            'aniversariantes_mes': aniversariantes_mes,
            'total_aniversariantes': total_aniversariantes,
        },
    )


def importar(request):
    if request.method == 'POST':
        form = HeadcountUploadForm(request.POST, request.FILES)
        if form.is_valid():
            mes = form.cleaned_data['mes']
            if isinstance(mes, datetime):
                mes = mes.date()
            mes = mes.replace(day=1)

            importacao = HeadcountImport.objects.create(
                mes=mes,
                arquivo=form.cleaned_data['headcount_file'],
            )

            try:
                import_headcount(importacao)
            except ValueError as exc:
                importacao.delete()
                messages.error(request, str(exc))
                return render(request, 'headcount/importar.html', {'form': form})

            lista = BirthdayList.objects.create(
                mes=mes,
                imagem=form.cleaned_data['birthday_image'],
                texto_extraido=form.cleaned_data['birthday_text'],
                headcount=importacao,
            )
            process_birthday_list(lista)
            messages.success(request, 'Headcount e aniversariantes importados com sucesso.')
            return redirect('headcount:detalhe', pk=importacao.pk)
    else:
        form = HeadcountUploadForm(initial={'mes': timezone.localdate().strftime('%Y-%m')})

    return render(request, 'headcount/importar.html', {'form': form})


def painel(request):
    importacoes = HeadcountImport.objects.all()
    return render(request, 'headcount/painel.html', {'importacoes': importacoes})


def detalhe(request, pk):
    importacao = get_object_or_404(HeadcountImport, pk=pk)
    lista = importacao.listas_aniversariantes.first()
    comparativo = compare_with_previous_month(importacao)
    por_turno = importacao.membros.values('turno').annotate(total=Count('id')).order_by('-total')
    por_area = importacao.membros.values('area').annotate(total=Count('id')).order_by('-total')[:10]

    return render(
        request,
        'headcount/detalhe.html',
        {
            'importacao': importacao,
            'lista': lista,
            'comparativo': comparativo,
            'por_turno': por_turno,
            'por_area': por_area,
        },
    )


def exportar_excel(request, pk):
    importacao = get_object_or_404(HeadcountImport, pk=pk)
    lista = importacao.listas_aniversariantes.first()
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'Aniversariantes'

    sheet.merge_cells('A1:E1')
    sheet['A1'] = f'Aniversariantes e Headcount - {importacao.mes.strftime("%m/%Y")}'
    sheet['A1'].font = Font(bold=True, size=16, color='FFFFFF')
    sheet['A1'].fill = PatternFill('solid', fgColor='17212B')
    sheet['A1'].alignment = Alignment(horizontal='center')
    sheet.append(['Nome', 'Encontrado no headcount', 'Turno', 'Work group', 'Team'])

    header_fill = PatternFill('solid', fgColor='E8F3EF')
    for cell in sheet[2]:
        cell.font = Font(bold=True, color='17212B')
        cell.fill = header_fill

    if lista:
        for aniversariante in lista.nomes.select_related('membro'):
            membro = aniversariante.membro
            sheet.append(
                [
                    aniversariante.nome,
                    'Sim' if membro else 'Nao',
                    membro.turno if membro else '',
                    membro.work_group if membro else '',
                    membro.team if membro else '',
                ]
            )

    members_sheet = workbook.create_sheet('Headcount')
    members_sheet.append(['Nome', 'Turno', 'Work group', 'Team', 'Area'])
    for member in importacao.membros.all():
        members_sheet.append([member.nome, member.turno, member.work_group, member.team, member.area])

    for active_sheet in workbook.worksheets:
        for row in active_sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical='top', wrap_text=True)
        for index, width in enumerate([32, 18, 18, 26, 26], start=1):
            active_sheet.column_dimensions[get_column_letter(index)].width = width

    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    filename = f"headcount_aniversariantes_{importacao.mes.strftime('%Y_%m')}.xlsx"
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response

# Create your views here.
