from io import BytesIO

from django.contrib import messages
from django.db.models import Count, Q
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .forms import TarefaCreateForm, TarefaFollowUpForm
from .models import Tarefa


STATUS_FINAIS = [Tarefa.Status.CONCLUIDA, Tarefa.Status.CANCELADA]


def index(request):
    hoje = timezone.localdate()
    abertas = Tarefa.objects.exclude(status__in=STATUS_FINAIS)
    mes = Tarefa.objects.filter(criado_em__year=hoje.year, criado_em__month=hoje.month)

    return render(
        request,
        'tarefas/index.html',
        {
            'abertas': abertas.count(),
            'vencidas': abertas.filter(prazo__lt=hoje).count(),
            'concluidas_mes': mes.filter(status=Tarefa.Status.CONCLUIDA).count(),
            'por_prioridade': abertas.values('prioridade').annotate(total=Count('id')).order_by('-total'),
        },
    )


def criar(request):
    if request.method == 'POST':
        form = TarefaCreateForm(request.POST)
        if form.is_valid():
            tarefa = form.save()
            messages.success(request, 'Tarefa de follow-up cadastrada com sucesso.')
            return redirect('tarefas:detalhe', pk=tarefa.pk)
    else:
        form = TarefaCreateForm()

    return render(request, 'tarefas/form.html', {'form': form, 'titulo': 'Nova tarefa', 'acao': 'Cadastrar tarefa'})


def painel(request):
    status = request.GET.get('status', '')
    prioridade = request.GET.get('prioridade', '')
    busca = request.GET.get('busca', '').strip()
    mes = request.GET.get('mes', '')

    tarefas = Tarefa.objects.all()

    if status:
        tarefas = tarefas.filter(status=status)

    if prioridade:
        tarefas = tarefas.filter(prioridade=prioridade)

    if busca:
        tarefas = tarefas.filter(
            Q(titulo__icontains=busca)
            | Q(descricao__icontains=busca)
            | Q(responsavel__icontains=busca)
            | Q(area__icontains=busca)
            | Q(origem__icontains=busca)
            | Q(follow_up__icontains=busca)
        )

    if mes:
        try:
            ano, numero_mes = mes.split('-')
            tarefas = tarefas.filter(criado_em__year=ano, criado_em__month=numero_mes)
        except ValueError:
            messages.error(request, 'Filtro de mes invalido.')

    return render(
        request,
        'tarefas/painel.html',
        {
            'tarefas': tarefas,
            'status_choices': Tarefa.Status.choices,
            'prioridade_choices': Tarefa.Prioridade.choices,
            'status_atual': status,
            'prioridade_atual': prioridade,
            'busca': busca,
            'mes': mes,
        },
    )


def kanban(request):
    tarefas = Tarefa.objects.all()
    colunas = [
        {
            'status': Tarefa.Status.PENDENTE,
            'titulo': 'A fazer',
            'tarefas': tarefas.filter(status=Tarefa.Status.PENDENTE),
        },
        {
            'status': Tarefa.Status.EM_ANDAMENTO,
            'titulo': 'Em andamento',
            'tarefas': tarefas.filter(status=Tarefa.Status.EM_ANDAMENTO),
        },
        {
            'status': Tarefa.Status.CONCLUIDA,
            'titulo': 'Concluido',
            'tarefas': tarefas.filter(status=Tarefa.Status.CONCLUIDA),
        },
        {
            'status': Tarefa.Status.CANCELADA,
            'titulo': 'Cancelado',
            'tarefas': tarefas.filter(status=Tarefa.Status.CANCELADA),
        },
    ]

    hoje = timezone.localdate()
    abertas = tarefas.exclude(status__in=STATUS_FINAIS)

    return render(
        request,
        'tarefas/kanban.html',
        {
            'colunas': colunas,
            'abertas': abertas.count(),
            'vencidas': abertas.filter(prazo__lt=hoje).count(),
            'em_andamento': tarefas.filter(status=Tarefa.Status.EM_ANDAMENTO).count(),
            'concluidas_mes': tarefas.filter(
                status=Tarefa.Status.CONCLUIDA,
                concluido_em__year=hoje.year,
                concluido_em__month=hoje.month,
            ).count(),
        },
    )


def alterar_status(request, pk):
    tarefa = get_object_or_404(Tarefa, pk=pk)

    if request.method == 'POST':
        novo_status = request.POST.get('status')
        if novo_status in Tarefa.Status.values:
            tarefa.status = novo_status
            if novo_status == Tarefa.Status.EM_ANDAMENTO and not tarefa.follow_up:
                tarefa.follow_up = 'Movida para em andamento pelo Kanban administrativo.'
            tarefa.save()
            messages.success(request, 'Status da tarefa atualizado.')
        else:
            messages.error(request, 'Status invalido.')

    return redirect(request.POST.get('next') or 'tarefas:kanban')


def detalhe(request, pk):
    tarefa = get_object_or_404(Tarefa, pk=pk)
    return render(request, 'tarefas/detalhe.html', {'tarefa': tarefa})


def editar(request, pk):
    tarefa = get_object_or_404(Tarefa, pk=pk)

    if request.method == 'POST':
        form = TarefaFollowUpForm(request.POST, instance=tarefa)
        if form.is_valid():
            form.save()
            messages.success(request, 'Follow-up atualizado com sucesso.')
            return redirect('tarefas:painel')
    else:
        form = TarefaFollowUpForm(instance=tarefa)

    return render(request, 'tarefas/form.html', {'form': form, 'titulo': 'Atualizar follow-up', 'acao': 'Salvar follow-up'})


def concluir(request, pk):
    tarefa = get_object_or_404(Tarefa, pk=pk)

    if request.method == 'POST':
        tarefa.status = Tarefa.Status.CONCLUIDA
        if not tarefa.follow_up:
            tarefa.follow_up = 'Baixa realizada pelo painel flutuante.'
        tarefa.save()
        messages.success(request, 'Tarefa baixada com sucesso.')

    return redirect(request.POST.get('next') or 'tarefas:painel')


def exportar_excel(request):
    tarefas = Tarefa.objects.all().order_by('-criado_em')
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'Follow-ups'

    sheet.merge_cells('A1:K1')
    sheet['A1'] = 'Relatorio de Tarefas e Follow-ups'
    sheet['A1'].font = Font(bold=True, size=16, color='FFFFFF')
    sheet['A1'].fill = PatternFill('solid', fgColor='17212B')
    sheet['A1'].alignment = Alignment(horizontal='center')

    headers = [
        'ID',
        'Mes',
        'Titulo',
        'Responsavel',
        'Area',
        'Origem',
        'Prioridade',
        'Status',
        'Prazo',
        'Descricao',
        'Follow-up',
    ]
    sheet.append(headers)

    header_fill = PatternFill('solid', fgColor='E8F3EF')
    for cell in sheet[2]:
        cell.font = Font(bold=True, color='17212B')
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    for tarefa in tarefas:
        sheet.append(
            [
                tarefa.id,
                tarefa.mes_referencia,
                tarefa.titulo,
                tarefa.responsavel,
                tarefa.area,
                tarefa.origem,
                tarefa.get_prioridade_display(),
                tarefa.get_status_display(),
                tarefa.prazo.strftime('%d/%m/%Y') if tarefa.prazo else '',
                tarefa.descricao,
                tarefa.follow_up,
            ]
        )

    for row in sheet.iter_rows(min_row=3):
        for cell in row:
            cell.alignment = Alignment(vertical='top', wrap_text=True)

    widths = [8, 12, 34, 24, 18, 22, 14, 18, 14, 46, 46]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width

    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    filename = f"relatorio_followups_{timezone.localdate().strftime('%Y_%m_%d')}.xlsx"
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response
