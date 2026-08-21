from io import BytesIO

from django.contrib import messages
from django.db.models import Count, Q
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .forms import ChamadoTICreateForm, ChamadoTIFollowUpForm
from .models import ChamadoTI


def get_mes_atual_bounds():
    hoje = timezone.localdate()
    return hoje.year, hoje.month


def index(request):
    ano, mes = get_mes_atual_bounds()
    chamados_mes = ChamadoTI.objects.filter(criado_em__year=ano, criado_em__month=mes)
    total_mes = chamados_mes.count()
    atendidos_mes = chamados_mes.filter(status=ChamadoTI.Status.CONCLUIDO).count()
    pendentes = ChamadoTI.objects.exclude(
        status__in=[ChamadoTI.Status.CONCLUIDO, ChamadoTI.Status.CANCELADO]
    ).count()
    por_categoria = chamados_mes.values('categoria').annotate(total=Count('id')).order_by('-total')

    return render(
        request,
        'chamados_ti/index.html',
        {
            'total_mes': total_mes,
            'atendidos_mes': atendidos_mes,
            'pendentes': pendentes,
            'por_categoria': por_categoria,
        },
    )


def criar(request):
    if request.method == 'POST':
        form = ChamadoTICreateForm(request.POST)
        if form.is_valid():
            chamado = form.save()
            messages.success(request, 'Chamado de TI cadastrado com sucesso.')
            return redirect('chamados_ti:detalhe', pk=chamado.pk)
    else:
        form = ChamadoTICreateForm()

    return render(request, 'chamados_ti/form.html', {'form': form})


def painel(request):
    status = request.GET.get('status', '')
    categoria = request.GET.get('categoria', '')
    busca = request.GET.get('busca', '').strip()
    mes = request.GET.get('mes', '')

    chamados = ChamadoTI.objects.all()

    if status:
        chamados = chamados.filter(status=status)

    if categoria:
        chamados = chamados.filter(categoria=categoria)

    if busca:
        chamados = chamados.filter(
            Q(titulo__icontains=busca)
            | Q(solicitante__icontains=busca)
            | Q(setor__icontains=busca)
            | Q(descricao__icontains=busca)
            | Q(ticket_oficial__icontains=busca)
        )

    if mes:
        try:
            ano, numero_mes = mes.split('-')
            chamados = chamados.filter(criado_em__year=ano, criado_em__month=numero_mes)
        except ValueError:
            messages.error(request, 'Filtro de mes invalido.')

    return render(
        request,
        'chamados_ti/painel.html',
        {
            'chamados': chamados,
            'status_choices': ChamadoTI.Status.choices,
            'categoria_choices': ChamadoTI.Categoria.choices,
            'status_atual': status,
            'categoria_atual': categoria,
            'busca': busca,
            'mes': mes,
        },
    )


def metricas(request):
    mes = request.GET.get('mes', timezone.localdate().strftime('%Y-%m'))
    chamados = ChamadoTI.objects.all()

    try:
        ano, numero_mes = mes.split('-')
        chamados = chamados.filter(criado_em__year=ano, criado_em__month=numero_mes)
    except ValueError:
        messages.error(request, 'Filtro de mes invalido.')
        mes = timezone.localdate().strftime('%Y-%m')
        ano, numero_mes = mes.split('-')
        chamados = chamados.filter(criado_em__year=ano, criado_em__month=numero_mes)

    total = chamados.count()
    atendidos = chamados.filter(status=ChamadoTI.Status.CONCLUIDO).count()
    cancelados = chamados.filter(status=ChamadoTI.Status.CANCELADO).count()
    pendentes = chamados.exclude(status__in=[ChamadoTI.Status.CONCLUIDO, ChamadoTI.Status.CANCELADO]).count()
    por_categoria = chamados.values('categoria').annotate(total=Count('id')).order_by('-total')
    por_status = chamados.values('status').annotate(total=Count('id')).order_by('-total')

    return render(
        request,
        'chamados_ti/metricas.html',
        {
            'mes': mes,
            'total': total,
            'atendidos': atendidos,
            'cancelados': cancelados,
            'pendentes': pendentes,
            'por_categoria': por_categoria,
            'por_status': por_status,
        },
    )


def detalhe(request, pk):
    chamado = get_object_or_404(ChamadoTI, pk=pk)
    return render(request, 'chamados_ti/detalhe.html', {'chamado': chamado})


def editar(request, pk):
    chamado = get_object_or_404(ChamadoTI, pk=pk)

    if request.method == 'POST':
        form = ChamadoTIFollowUpForm(request.POST, instance=chamado)
        if form.is_valid():
            form.save()
            messages.success(request, 'Follow-up de TI atualizado com sucesso.')
            return redirect('chamados_ti:painel')
    else:
        form = ChamadoTIFollowUpForm(instance=chamado)

    return render(request, 'chamados_ti/editar.html', {'form': form, 'chamado': chamado})


def exportar_excel(request):
    chamados = ChamadoTI.objects.all().order_by('-criado_em')
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'Chamados TI'

    sheet.merge_cells('A1:L1')
    sheet['A1'] = 'Relatorio de Chamados de TI'
    sheet['A1'].font = Font(bold=True, size=16, color='FFFFFF')
    sheet['A1'].fill = PatternFill('solid', fgColor='1967D2')
    sheet['A1'].alignment = Alignment(horizontal='center')

    headers = [
        'ID',
        'Mes',
        'Titulo',
        'Solicitante',
        'Setor',
        'Categoria',
        'Prioridade',
        'Descricao',
        'Status',
        'Ticket oficial',
        'Solucao/Follow-up',
        'Concluido em',
    ]
    sheet.append(headers)

    header_fill = PatternFill('solid', fgColor='E8F3EF')
    for cell in sheet[2]:
        cell.font = Font(bold=True, color='17212B')
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    for chamado in chamados:
        sheet.append(
            [
                chamado.id,
                chamado.mes_referencia,
                chamado.titulo,
                chamado.solicitante,
                chamado.setor,
                chamado.get_categoria_display(),
                chamado.get_prioridade_display(),
                chamado.descricao,
                chamado.get_status_display(),
                chamado.ticket_oficial,
                chamado.solucao,
                timezone.localtime(chamado.concluido_em).strftime('%d/%m/%Y %H:%M') if chamado.concluido_em else '',
            ]
        )

    for row in sheet.iter_rows(min_row=3):
        for cell in row:
            cell.alignment = Alignment(vertical='top', wrap_text=True)

    widths = [8, 12, 30, 24, 18, 24, 14, 45, 18, 18, 42, 20]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width

    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    filename = f"relatorio_chamados_ti_{timezone.localdate().strftime('%Y_%m_%d')}.xlsx"
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response

# Create your views here.
