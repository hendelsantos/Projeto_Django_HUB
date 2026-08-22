from io import BytesIO

import qrcode
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db.models import Q
from django.http import HttpResponse
from django.shortcuts import render
from django.urls import reverse
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from chamados_ti.models import ChamadoTI
from headcount.models import BirthdayName, HeadcountMember
from roupeiro.models import Armario
from zeladoria.models import ChamadoZeladoria

from .report_sources import build_consolidated_report


PHOTO_CLOUD_URL = 'https://photo-cloud-1.onrender.com/'


def get_month_filter(request):
    mes = request.GET.get('mes', timezone.localdate().strftime('%Y-%m'))

    try:
        ano, numero_mes = mes.split('-')
        return mes, int(ano), int(numero_mes)
    except (TypeError, ValueError):
        messages.error(request, 'Filtro de mes invalido.')
        mes = timezone.localdate().strftime('%Y-%m')
        ano, numero_mes = mes.split('-')
        return mes, int(ano), int(numero_mes)


def home(request):
    tools = [
        {
            'name': 'Scanner para Excel',
            'description': 'Suba um PDF escaneado da pintura e gere uma planilha Excel com os dados extraidos.',
            'details': 'Use quando receber arquivos de scanner e precisar transformar o conteudo em uma planilha para conferencia ou tratamento.',
            'url_name': 'extrator_scanner:index',
            'status': 'Primeiro app',
            'icon': 'PDF',
        },
        {
            'name': 'Zeladoria Predial',
            'description': 'Cadastre melhorias do predio, acompanhe tickets oficiais e exporte chamados por mes.',
            'details': 'Ideal para registrar pedidos da equipe sobre estrutura predial, inserir foto, fazer follow-up e controlar o que ainda esta aberto.',
            'url_name': 'zeladoria:index',
            'status': 'Novo app',
            'icon': 'ZEL',
        },
        {
            'name': 'Chamados de TI',
            'description': 'Controle atendimentos de TI, contas, sistemas e equipamentos com metricas mensais.',
            'details': 'Organiza demandas como abertura de conta, acesso a sistemas, verificacao de equipamentos e relatorio de atendimentos do mes.',
            'url_name': 'chamados_ti:index',
            'status': 'Gestao',
            'icon': 'TI',
        },
        {
            'name': 'Gestao de Roupeiro',
            'description': 'Controle armarios, usuarios, turnos e tamanhos de uniforme em um painel unico.',
            'details': 'Use para saber quais armarios estao livres, ocupados, em manutencao e quais pessoas possuem armario cadastrado.',
            'url_name': 'roupeiro:index',
            'status': 'Novo app',
            'icon': 'ARM',
        },
        {
            'name': 'Headcount e Aniversariantes',
            'description': 'Importe o headcount mensal, cruze com aniversariantes e acompanhe entradas, saidas, turnos e areas.',
            'details': 'Use para gerar listas mensais de aniversariantes por turno e area, alem de comparar evolucao do headcount da pintura.',
            'url_name': 'headcount:index',
            'status': 'Metricas',
            'icon': 'HC',
        },
    ]

    if request.user.is_authenticated:
        tools.append(
            {
                'name': 'PhotoCloud',
                'description': 'Envie fotos pelo celular para uma galeria temporaria usando link direto ou QR Code.',
                'details': 'Ideal para registrar evidencias visuais rapidamente: abra o QR Code no celular, informe nome e album, selecione as fotos e envie.',
                'url_name': 'hub:photocloud',
                'status': 'Fotos',
                'icon': 'FOTO',
            }
        )

    return render(request, 'hub/home.html', {'tools': tools})


@login_required(login_url='admin:login')
def photocloud(request):
    return render(
        request,
        'hub/photocloud.html',
        {
            'photo_cloud_url': PHOTO_CLOUD_URL,
        },
    )


@login_required(login_url='admin:login')
def photocloud_qrcode(request):
    image = qrcode.make(PHOTO_CLOUD_URL)
    output = BytesIO()
    image.save(output, format='PNG')
    output.seek(0)

    return HttpResponse(output.getvalue(), content_type='image/png')


def buscar(request):
    query = request.GET.get('q', '').strip()
    resultados = []

    if query:
        resultados.extend(buscar_roupeiro(query))
        resultados.extend(buscar_ti(query))
        resultados.extend(buscar_zeladoria(query))
        resultados.extend(buscar_headcount(query))

    return render(
        request,
        'hub/busca.html',
        {
            'query': query,
            'resultados': resultados,
            'total_resultados': len(resultados),
        },
    )


def buscar_roupeiro(query):
    filtros = Q(usuario__icontains=query) | Q(observacoes__icontains=query)
    if query.isdigit():
        numero = int(query)
        filtros |= (
            Q(numero=numero)
            | Q(tamanho_camisa_numero=numero)
            | Q(tamanho_calca_numero=numero)
            | Q(tamanho_macacao_numero=numero)
        )

    return [
        {
            'modulo': 'Roupeiro',
            'titulo': f'Armario #{armario.numero}',
            'descricao': f'{armario.usuario or "Sem usuario"} - {armario.get_status_display()}',
            'extra': armario.get_turno_display() if armario.turno else 'Sem turno',
            'url': reverse('roupeiro:detalhe', args=[armario.pk]),
        }
        for armario in Armario.objects.filter(filtros)[:10]
    ]


def buscar_ti(query):
    chamados = ChamadoTI.objects.filter(
        Q(titulo__icontains=query)
        | Q(solicitante__icontains=query)
        | Q(setor__icontains=query)
        | Q(descricao__icontains=query)
        | Q(ticket_oficial__icontains=query)
        | Q(solucao__icontains=query)
    )[:10]

    return [
        {
            'modulo': 'TI',
            'titulo': chamado.titulo,
            'descricao': f'{chamado.solicitante} - {chamado.get_status_display()}',
            'extra': chamado.get_categoria_display(),
            'url': reverse('chamados_ti:detalhe', args=[chamado.pk]),
        }
        for chamado in chamados
    ]


def buscar_zeladoria(query):
    chamados = ChamadoZeladoria.objects.filter(
        Q(titulo__icontains=query)
        | Q(solicitante__icontains=query)
        | Q(local__icontains=query)
        | Q(descricao__icontains=query)
        | Q(ticket_oficial__icontains=query)
        | Q(observacoes__icontains=query)
    )[:10]

    return [
        {
            'modulo': 'Zeladoria',
            'titulo': chamado.titulo,
            'descricao': f'{chamado.solicitante} - {chamado.get_status_display()}',
            'extra': chamado.local,
            'url': reverse('zeladoria:detalhe', args=[chamado.pk]),
        }
        for chamado in chamados
    ]


def buscar_headcount(query):
    membros = HeadcountMember.objects.filter(
        Q(nome__icontains=query)
        | Q(turno__icontains=query)
        | Q(work_group__icontains=query)
        | Q(team__icontains=query)
        | Q(area__icontains=query)
    ).select_related('importacao')[:10]
    aniversariantes = BirthdayName.objects.filter(nome__icontains=query).select_related('lista', 'membro')[:10]

    resultados = [
        {
            'modulo': 'Headcount',
            'titulo': membro.nome,
            'descricao': f'{membro.turno or "Sem turno"} - {membro.area or "Sem area"}',
            'extra': membro.importacao.mes.strftime('%m/%Y'),
            'url': reverse('headcount:detalhe', args=[membro.importacao.pk]),
        }
        for membro in membros
    ]

    resultados.extend(
        {
            'modulo': 'Aniversariantes',
            'titulo': aniversariante.nome,
            'descricao': 'Encontrado no headcount' if aniversariante.membro else 'Nao encontrado no headcount',
            'extra': aniversariante.lista.mes.strftime('%m/%Y'),
            'url': reverse('headcount:detalhe', args=[aniversariante.lista.headcount.pk]),
        }
        for aniversariante in aniversariantes
    )
    return resultados


def relatorios(request):
    mes, ano, numero_mes = get_month_filter(request)
    report = build_consolidated_report(ano, numero_mes)
    ti_source = next((source for source in report['sources'] if source['slug'] == 'ti'), None)
    zeladoria_source = next((source for source in report['sources'] if source['slug'] == 'zeladoria'), None)

    context = {
        'mes': mes,
        'total_geral': report['total'],
        'concluidos': report['concluidos'],
        'pendentes': report['pendentes'],
        'cancelados': report['cancelados'],
        'sem_ticket': report['sem_ticket'],
        'taxa_conclusao': report['taxa_conclusao'],
        'fontes_relatorio': report['sources'],
        'areas': report['consolidated_sources'],
        'ti_por_status': ti_source['por_status'] if ti_source else [],
        'ti_por_categoria': ti_source['por_categoria'] if ti_source else [],
        'zeladoria_por_status': zeladoria_source['por_status'] if zeladoria_source else [],
        'itens_recentes': report['items'][:12],
    }

    return render(request, 'hub/relatorios.html', context)


def exportar_relatorio_geral(request):
    mes, ano, numero_mes = get_month_filter(request)
    report = build_consolidated_report(ano, numero_mes)
    itens = report['items']

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'Relatorio Geral'

    sheet.merge_cells('A1:J1')
    sheet['A1'] = f'Relatorio Geral Ferramentas digitais Paint Shop - {mes}'
    sheet['A1'].font = Font(bold=True, size=16, color='FFFFFF')
    sheet['A1'].fill = PatternFill('solid', fgColor='17212B')
    sheet['A1'].alignment = Alignment(horizontal='center')

    headers = [
        'Area',
        'Data',
        'Titulo',
        'Solicitante',
        'Referencia',
        'Status',
        'Ticket oficial',
        'Descricao',
        'Follow-up',
        'Mes',
    ]
    sheet.append(headers)

    header_fill = PatternFill('solid', fgColor='E8F3EF')
    for cell in sheet[2]:
        cell.font = Font(bold=True, color='17212B')
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    for item in itens:
        sheet.append(
            [
                item['area'],
                timezone.localtime(item['data']).strftime('%d/%m/%Y %H:%M'),
                item['titulo'],
                item['solicitante'],
                item['referencia'],
                item['status'],
                item['ticket'],
                item['detalhe'],
                item['follow_up'],
                mes,
            ]
        )

    for row in sheet.iter_rows(min_row=3):
        for cell in row:
            cell.alignment = Alignment(vertical='top', wrap_text=True)

    widths = [16, 18, 32, 24, 28, 20, 18, 46, 42, 12]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width

    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    filename = f"relatorio_geral_paint_hub_{mes.replace('-', '_')}.xlsx"
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response
