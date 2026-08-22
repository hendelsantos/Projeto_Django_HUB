import re
import unicodedata

from openpyxl import load_workbook

from .models import BirthdayName, HeadcountImport, HeadcountMember


NAME_HEADERS = ('nome', 'name', 'employee name', 'member', 'membro', 'colaborador')
SHIFT_HEADERS = ('turno', 'shift')
WORK_GROUP_HEADERS = ('work group', 'workgroup', 'grupo de trabalho')
TEAM_HEADERS = ('team', 'tea', 'time', 'equipe')


def normalize_text(value):
    value = str(value or '').strip().lower()
    value = unicodedata.normalize('NFKD', value)
    value = ''.join(char for char in value if not unicodedata.combining(char))
    value = re.sub(r'[^a-z0-9 ]+', ' ', value)
    return re.sub(r'\s+', ' ', value).strip()


def normalize_name(value):
    return normalize_text(value)


def find_column(headers, candidates):
    normalized_headers = [normalize_text(header) for header in headers]
    for candidate in candidates:
        candidate = normalize_text(candidate)
        if candidate in normalized_headers:
            return normalized_headers.index(candidate)
    for index, header in enumerate(normalized_headers):
        if any(candidate in header for candidate in candidates):
            return index
    return None


def import_headcount(importacao):
    workbook = load_workbook(importacao.arquivo.path, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))

    if not rows:
        return 0

    headers = [str(value or '') for value in rows[0]]
    name_index = find_column(headers, NAME_HEADERS)
    shift_index = find_column(headers, SHIFT_HEADERS)
    work_group_index = find_column(headers, WORK_GROUP_HEADERS)
    team_index = find_column(headers, TEAM_HEADERS)

    if name_index is None:
        raise ValueError('Nao encontrei uma coluna de nome no arquivo. Use uma coluna chamada Nome ou Name.')

    HeadcountMember.objects.filter(importacao=importacao).delete()
    members = []

    for row in rows[1:]:
        nome = row[name_index] if name_index < len(row) else ''
        nome_normalizado = normalize_name(nome)
        if not nome_normalizado:
            continue

        work_group = row[work_group_index] if work_group_index is not None and work_group_index < len(row) else ''
        team = row[team_index] if team_index is not None and team_index < len(row) else ''
        area = ' / '.join(str(value).strip() for value in [work_group, team] if value)

        members.append(
            HeadcountMember(
                importacao=importacao,
                nome=str(nome).strip(),
                nome_normalizado=nome_normalizado,
                turno=str(row[shift_index]).strip() if shift_index is not None and shift_index < len(row) and row[shift_index] else '',
                work_group=str(work_group).strip() if work_group else '',
                team=str(team).strip() if team else '',
                area=area,
            )
        )

    HeadcountMember.objects.bulk_create(members)
    importacao.total_membros = len(members)
    importacao.save(update_fields=['total_membros'])
    return len(members)


def parse_birthday_names(text):
    names = []
    for line in (text or '').splitlines():
        cleaned = re.sub(r'\d{1,2}[/-]\d{1,2}([/-]\d{2,4})?', '', line)
        cleaned = re.sub(r'\b(janeiro|fevereiro|marco|março|abril|maio|junho|julho|agosto|setembro|outubro|novembro|dezembro)\b', '', cleaned, flags=re.I)
        cleaned = re.sub(r'[^A-Za-zÀ-ÿ ]+', ' ', cleaned)
        cleaned = re.sub(r'\s+', ' ', cleaned).strip()
        if len(cleaned.split()) >= 2:
            names.append(cleaned)
    return names


def process_birthday_list(lista):
    BirthdayName.objects.filter(lista=lista).delete()
    members_by_name = {
        member.nome_normalizado: member
        for member in lista.headcount.membros.all()
    }

    birthday_names = []
    for nome in parse_birthday_names(lista.texto_extraido):
        normalized = normalize_name(nome)
        birthday_names.append(
            BirthdayName(
                lista=lista,
                nome=nome,
                nome_normalizado=normalized,
                membro=members_by_name.get(normalized),
            )
        )

    BirthdayName.objects.bulk_create(birthday_names)
    return len(birthday_names)


def compare_with_previous_month(importacao):
    previous = (
        HeadcountImport.objects.filter(mes__lt=importacao.mes)
        .order_by('-mes', '-criado_em')
        .first()
    )
    current_names = set(importacao.membros.values_list('nome_normalizado', flat=True))

    if previous is None:
        return {'entradas': current_names, 'saidas': set(), 'previous': None}

    previous_names = set(previous.membros.values_list('nome_normalizado', flat=True))
    return {
        'entradas': current_names - previous_names,
        'saidas': previous_names - current_names,
        'previous': previous,
    }
