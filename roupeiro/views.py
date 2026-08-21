from io import BytesIO

from django.contrib import messages
from django.db.models import Count, Max, Q
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .forms import ArmarioForm
from .models import Armario


def index(request):
    total_armarios = Armario.objects.count()
    ocupados = Armario.objects.filter(status=Armario.Status.OCUPADO).count()
    livres = Armario.objects.filter(status=Armario.Status.LIVRE).count()
    manutencao = Armario.objects.filter(status=Armario.Status.MANUTENCAO).count()

    return render(
        request,
        'roupeiro/index.html',
        {
            'total_armarios': total_armarios,
            'ocupados': ocupados,
            'livres': livres,
            'manutencao': manutencao,
        },
    )


def criar(request):
    if request.method == 'POST':
        form = ArmarioForm(request.POST)
        if form.is_valid():
            armario = form.save()
            messages.success(request, 'Armario cadastrado com sucesso.')
            return redirect('roupeiro:detalhe', pk=armario.pk)
    else:
        proximo_numero = (Armario.objects.aggregate(maior_numero=Max('numero'))['maior_numero'] or 0) + 1
        form = ArmarioForm(initial={'numero': proximo_numero})

    return render(request, 'roupeiro/form.html', {'form': form, 'titulo': 'Cadastrar armario'})


def painel(request):
    status = request.GET.get('status', '')
    turno = request.GET.get('turno', '')
    tamanho = request.GET.get('tamanho', '')
    busca = request.GET.get('busca', '').strip()

    armarios = Armario.objects.all()

    if status:
        armarios = armarios.filter(status=status)

    if turno:
        armarios = armarios.filter(turno=turno)

    if tamanho:
        armarios = armarios.filter(
            Q(tamanho_camisa=tamanho)
            | Q(tamanho_calca=tamanho)
            | Q(tamanho_macacao=tamanho)
        )

    if busca:
        numero_busca = int(busca) if busca.isdigit() else None
        filtros_busca = Q(usuario__icontains=busca) | Q(observacoes__icontains=busca)
        if numero_busca is not None:
            filtros_busca |= (
                Q(numero=numero_busca)
                | Q(tamanho_camisa_numero=numero_busca)
                | Q(tamanho_calca_numero=numero_busca)
                | Q(tamanho_macacao_numero=numero_busca)
            )

        armarios = armarios.filter(filtros_busca)

    resumo_status = Armario.objects.values('status').annotate(total=Count('id')).order_by('-total')

    return render(
        request,
        'roupeiro/painel.html',
        {
            'armarios': armarios,
            'status_choices': Armario.Status.choices,
            'turno_choices': Armario.Turno.choices,
            'tamanho_choices': Armario.TamanhoRoupa.choices,
            'status_atual': status,
            'turno_atual': turno,
            'tamanho_atual': tamanho,
            'busca': busca,
            'resumo_status': resumo_status,
        },
    )


def mapa(request):
    status = request.GET.get('status', '')
    busca = request.GET.get('busca', '').strip()

    armarios = Armario.objects.all()

    if status:
        armarios = armarios.filter(status=status)

    if busca:
        numero_busca = int(busca) if busca.isdigit() else None
        filtros_busca = Q(usuario__icontains=busca) | Q(observacoes__icontains=busca)
        if numero_busca is not None:
            filtros_busca |= Q(numero=numero_busca)
        armarios = armarios.filter(filtros_busca)

    total_armarios = Armario.objects.count()
    ocupados = Armario.objects.filter(status=Armario.Status.OCUPADO).count()
    livres = Armario.objects.filter(status=Armario.Status.LIVRE).count()
    manutencao = Armario.objects.filter(status=Armario.Status.MANUTENCAO).count()

    return render(
        request,
        'roupeiro/mapa.html',
        {
            'armarios': armarios,
            'status_choices': Armario.Status.choices,
            'status_atual': status,
            'busca': busca,
            'total_armarios': total_armarios,
            'ocupados': ocupados,
            'livres': livres,
            'manutencao': manutencao,
        },
    )


def detalhe(request, pk):
    armario = get_object_or_404(Armario, pk=pk)
    return render(request, 'roupeiro/detalhe.html', {'armario': armario})


def editar(request, pk):
    armario = get_object_or_404(Armario, pk=pk)

    if request.method == 'POST':
        form = ArmarioForm(request.POST, instance=armario)
        if form.is_valid():
            form.save()
            messages.success(request, 'Armario atualizado com sucesso.')
            return redirect('roupeiro:detalhe', pk=armario.pk)
    else:
        form = ArmarioForm(instance=armario)

    return render(request, 'roupeiro/form.html', {'form': form, 'titulo': f'Editar armario #{armario.numero}', 'armario': armario})


def liberar(request, pk):
    armario = get_object_or_404(Armario, pk=pk)

    if request.method != 'POST':
        return redirect('roupeiro:detalhe', pk=armario.pk)

    armario.usuario = ''
    armario.turno = ''
    armario.tamanho_camisa = ''
    armario.tamanho_camisa_numero = None
    armario.tamanho_calca = ''
    armario.tamanho_calca_numero = None
    armario.tamanho_macacao = ''
    armario.tamanho_macacao_numero = None
    armario.status = Armario.Status.LIVRE
    armario.observacoes = f'{armario.observacoes}\nArmario liberado em {timezone.localtime().strftime("%d/%m/%Y %H:%M")}.'.strip()
    armario.save()
    messages.success(request, 'Armario liberado com sucesso.')
    return redirect('roupeiro:detalhe', pk=armario.pk)


def exportar_excel(request):
    armarios = Armario.objects.all()
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'Armarios'

    sheet.merge_cells('A1:M1')
    sheet['A1'] = 'Relatorio de Armarios do Roupeiro'
    sheet['A1'].font = Font(bold=True, size=16, color='FFFFFF')
    sheet['A1'].fill = PatternFill('solid', fgColor='1967D2')
    sheet['A1'].alignment = Alignment(horizontal='center')

    headers = [
        'Numero',
        'Status',
        'Usuario',
        'Turno',
        'Camisa',
        'Camisa numero',
        'Calca',
        'Calca numero',
        'Macacao',
        'Macacao numero',
        'Observacoes',
        'Criado em',
        'Atualizado em',
    ]
    sheet.append(headers)

    header_fill = PatternFill('solid', fgColor='E8F3EF')
    for cell in sheet[2]:
        cell.font = Font(bold=True, color='17212B')
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    for armario in armarios:
        sheet.append(
            [
                armario.numero,
                armario.get_status_display(),
                armario.usuario,
                armario.get_turno_display() if armario.turno else '',
                armario.get_tamanho_camisa_display() if armario.tamanho_camisa else '',
                armario.tamanho_camisa_numero or '',
                armario.get_tamanho_calca_display() if armario.tamanho_calca else '',
                armario.tamanho_calca_numero or '',
                armario.get_tamanho_macacao_display() if armario.tamanho_macacao else '',
                armario.tamanho_macacao_numero or '',
                armario.observacoes,
                timezone.localtime(armario.criado_em).strftime('%d/%m/%Y %H:%M'),
                timezone.localtime(armario.atualizado_em).strftime('%d/%m/%Y %H:%M'),
            ]
        )

    for row in sheet.iter_rows(min_row=3):
        for cell in row:
            cell.alignment = Alignment(vertical='top', wrap_text=True)

    widths = [10, 16, 28, 18, 12, 15, 12, 14, 12, 16, 42, 18, 18]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width

    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    filename = f"relatorio_armarios_roupeiro_{timezone.localdate().strftime('%Y_%m_%d')}.xlsx"
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response

# Create your views here.
