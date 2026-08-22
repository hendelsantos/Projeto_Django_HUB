from calendar import monthrange
from datetime import date, timedelta
from io import BytesIO

from django.contrib import messages
from django.db.models import Count, Q
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .forms import TreinamentoForm
from .models import TreinamentoSeguranca
from .services import fill_participants_from_pdf, sync_participants


def index(request):
    hoje = timezone.localdate()
    treinamentos_mes = TreinamentoSeguranca.objects.filter(data__year=hoje.year, data__month=hoje.month)
    total_participantes_mes = sum(item.total_participantes for item in treinamentos_mes)
    vencidos = TreinamentoSeguranca.objects.filter(validade__lt=hoje).count()

    return render(
        request,
        'treinamentos/index.html',
        {
            'treinamentos_mes': treinamentos_mes.count(),
            'participantes_mes': total_participantes_mes,
            'vencidos': vencidos,
            'por_empresa': treinamentos_mes.values('empresa').annotate(total=Count('id')).order_by('-total')[:6],
        },
    )


def criar(request):
    if request.method == 'POST':
        form = TreinamentoForm(request.POST, request.FILES)
        if form.is_valid():
            treinamento = form.save()
            extracted_from_pdf = fill_participants_from_pdf(treinamento)
            sync_participants(treinamento)
            if extracted_from_pdf:
                messages.success(request, 'Treinamento cadastrado e participantes extraidos do PDF com sucesso.')
            elif treinamento.documento and not treinamento.texto_participantes.strip():
                messages.warning(request, 'Treinamento cadastrado. O PDF parece ser scanner/imagem; cole os nomes para gerar o Excel estruturado.')
            else:
                messages.success(request, 'Treinamento cadastrado com sucesso.')
            return redirect('treinamentos:detalhe', pk=treinamento.pk)
    else:
        data_inicial = request.GET.get('data') or timezone.localdate().isoformat()
        form = TreinamentoForm(
            initial={
                'data': data_inicial,
                'area': 'Pintura',
                'status': TreinamentoSeguranca.Status.AGENDADO,
            }
        )

    return render(
        request,
        'treinamentos/form.html',
        {'form': form, 'titulo': 'Novo treinamento', 'acao': 'Cadastrar treinamento'},
    )


def painel(request):
    busca = request.GET.get('busca', '').strip()
    empresa = request.GET.get('empresa', '').strip()
    categoria = request.GET.get('categoria', '')
    mes = request.GET.get('mes', '')

    treinamentos = TreinamentoSeguranca.objects.all()

    if busca:
        treinamentos = treinamentos.filter(
            Q(titulo__icontains=busca)
            | Q(instrutor__icontains=busca)
            | Q(area__icontains=busca)
            | Q(observacoes__icontains=busca)
            | Q(participantes__nome__icontains=busca)
        ).distinct()

    if empresa:
        treinamentos = treinamentos.filter(empresa__icontains=empresa)

    if categoria:
        treinamentos = treinamentos.filter(categoria=categoria)

    if mes:
        try:
            ano, numero_mes = mes.split('-')
            treinamentos = treinamentos.filter(data__year=ano, data__month=numero_mes)
        except ValueError:
            messages.error(request, 'Filtro de mes invalido.')

    return render(
        request,
        'treinamentos/painel.html',
        {
            'treinamentos': treinamentos,
            'busca': busca,
            'empresa': empresa,
            'categoria_atual': categoria,
            'mes': mes,
            'categoria_choices': TreinamentoSeguranca.Categoria.choices,
        },
    )


def calendario(request):
    mes = request.GET.get('mes', timezone.localdate().strftime('%Y-%m'))
    try:
        ano, numero_mes = [int(part) for part in mes.split('-')]
    except (TypeError, ValueError):
        messages.error(request, 'Filtro de mes invalido.')
        hoje = timezone.localdate()
        ano, numero_mes = hoje.year, hoje.month
        mes = hoje.strftime('%Y-%m')

    primeiro_dia = date(ano, numero_mes, 1)
    total_dias = monthrange(ano, numero_mes)[1]
    inicio_grade = primeiro_dia - timedelta(days=primeiro_dia.weekday())
    dias = [inicio_grade + timedelta(days=index) for index in range(42)]
    treinamentos = TreinamentoSeguranca.objects.filter(data__year=ano, data__month=numero_mes).order_by('data', 'hora_inicio')

    agenda_por_dia = {}
    for treinamento in treinamentos:
        agenda_por_dia.setdefault(treinamento.data, []).append(treinamento)

    mes_anterior_data = (primeiro_dia - timedelta(days=1)).replace(day=1)
    proximo_mes_numero = numero_mes + 1
    proximo_ano = ano
    if proximo_mes_numero == 13:
        proximo_mes_numero = 1
        proximo_ano += 1
    proximo_mes_data = date(proximo_ano, proximo_mes_numero, 1)

    return render(
        request,
        'treinamentos/calendario.html',
        {
            'mes': mes,
            'mes_label': primeiro_dia.strftime('%m/%Y'),
            'mes_anterior': mes_anterior_data.strftime('%Y-%m'),
            'proximo_mes': proximo_mes_data.strftime('%Y-%m'),
            'dias': [
                {
                    'data': dia,
                    'fora_mes': dia.month != numero_mes,
                    'treinamentos': agenda_por_dia.get(dia, []),
                }
                for dia in dias
            ],
            'total_agendados': treinamentos.filter(status=TreinamentoSeguranca.Status.AGENDADO).count(),
            'total_realizados': treinamentos.filter(status=TreinamentoSeguranca.Status.REALIZADO).count(),
            'total_cancelados': treinamentos.filter(status=TreinamentoSeguranca.Status.CANCELADO).count(),
        },
    )


def detalhe(request, pk):
    treinamento = get_object_or_404(TreinamentoSeguranca, pk=pk)
    por_empresa = treinamento.participantes.values('empresa').annotate(total=Count('id')).order_by('-total')
    por_area = treinamento.participantes.values('area').annotate(total=Count('id')).order_by('-total')
    por_turno = treinamento.participantes.values('turno').annotate(total=Count('id')).order_by('-total')

    return render(
        request,
        'treinamentos/detalhe.html',
        {
            'treinamento': treinamento,
            'por_empresa': por_empresa,
            'por_area': por_area,
            'por_turno': por_turno,
        },
    )


def editar(request, pk):
    treinamento = get_object_or_404(TreinamentoSeguranca, pk=pk)

    if request.method == 'POST':
        form = TreinamentoForm(request.POST, request.FILES, instance=treinamento)
        if form.is_valid():
            treinamento = form.save()
            extracted_from_pdf = fill_participants_from_pdf(treinamento)
            sync_participants(treinamento)
            if extracted_from_pdf:
                messages.success(request, 'Treinamento atualizado e participantes extraidos do PDF com sucesso.')
            elif treinamento.documento and not treinamento.texto_participantes.strip():
                messages.warning(request, 'Treinamento atualizado. O PDF parece ser scanner/imagem; cole os nomes para gerar o Excel estruturado.')
            else:
                messages.success(request, 'Treinamento atualizado com sucesso.')
            return redirect('treinamentos:detalhe', pk=treinamento.pk)
    else:
        form = TreinamentoForm(instance=treinamento)

    return render(
        request,
        'treinamentos/form.html',
        {'form': form, 'titulo': 'Editar treinamento', 'acao': 'Salvar treinamento'},
    )


def exportar_treinamento_excel(request, pk):
    treinamento = get_object_or_404(TreinamentoSeguranca.objects.prefetch_related('participantes'), pk=pk)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'Participantes'

    sheet.merge_cells('A1:H1')
    sheet['A1'] = f'{treinamento.titulo} - {treinamento.data.strftime("%d/%m/%Y")}'
    sheet['A1'].font = Font(bold=True, size=16, color='FFFFFF')
    sheet['A1'].fill = PatternFill('solid', fgColor='17212B')
    sheet['A1'].alignment = Alignment(horizontal='center')

    sheet.append(['Nome', 'Matricula', 'Empresa', 'Turno', 'Area', 'Treinamento', 'Data', 'Horario'])
    header_fill = PatternFill('solid', fgColor='E8F3EF')
    for cell in sheet[2]:
        cell.font = Font(bold=True, color='17212B')
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    for participante in treinamento.participantes.all():
        sheet.append(
            [
                participante.nome,
                participante.matricula,
                participante.empresa,
                participante.turno,
                participante.area,
                treinamento.titulo,
                treinamento.data.strftime('%d/%m/%Y'),
                treinamento.horario,
            ]
        )

    for row in sheet.iter_rows(min_row=3):
        for cell in row:
            cell.alignment = Alignment(vertical='top', wrap_text=True)

    for index, width in enumerate([34, 16, 24, 14, 24, 36, 14, 18], start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width

    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    filename = f"participantes_treinamento_{treinamento.pk}_{treinamento.data.strftime('%Y_%m_%d')}.xlsx"
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


def exportar_excel(request):
    treinamentos = TreinamentoSeguranca.objects.prefetch_related('participantes').all()
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'Treinamentos'

    sheet.merge_cells('A1:L1')
    sheet['A1'] = 'Relatorio de Treinamentos de Seguranca da Pintura'
    sheet['A1'].font = Font(bold=True, size=16, color='FFFFFF')
    sheet['A1'].fill = PatternFill('solid', fgColor='17212B')
    sheet['A1'].alignment = Alignment(horizontal='center')

    headers = [
        'ID',
        'Data',
        'Horario',
        'Titulo',
        'Categoria',
        'Status',
        'Empresa',
        'Area',
        'Instrutor',
        'Carga horaria',
        'Validade',
        'Participantes',
    ]
    sheet.append(headers)

    header_fill = PatternFill('solid', fgColor='E8F3EF')
    for cell in sheet[2]:
        cell.font = Font(bold=True, color='17212B')
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    for treinamento in treinamentos:
        sheet.append(
            [
                treinamento.id,
                treinamento.data.strftime('%d/%m/%Y'),
                treinamento.horario,
                treinamento.titulo,
                treinamento.get_categoria_display(),
                treinamento.get_status_display(),
                treinamento.empresa,
                treinamento.area,
                treinamento.instrutor,
                treinamento.carga_horaria,
                treinamento.validade.strftime('%d/%m/%Y') if treinamento.validade else '',
                treinamento.total_participantes,
            ]
        )

    participantes_sheet = workbook.create_sheet('Participantes')
    participantes_sheet.append(['Treinamento', 'Data', 'Horario', 'Nome', 'Matricula', 'Empresa', 'Turno', 'Area'])
    for treinamento in treinamentos:
        for participante in treinamento.participantes.all():
            participantes_sheet.append(
                [
                    treinamento.titulo,
                    treinamento.data.strftime('%d/%m/%Y'),
                    treinamento.horario,
                    participante.nome,
                    participante.matricula,
                    participante.empresa,
                    participante.turno,
                    participante.area,
                ]
            )

    for active_sheet in workbook.worksheets:
        for row in active_sheet.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical='top', wrap_text=True)
        for index, width in enumerate([10, 14, 18, 36, 20, 18, 24, 20, 24, 16, 16, 16, 18], start=1):
            active_sheet.column_dimensions[get_column_letter(index)].width = width

    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    filename = f"relatorio_treinamentos_seguranca_{timezone.localdate().strftime('%Y_%m_%d')}.xlsx"
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response
